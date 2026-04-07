from django.shortcuts import get_object_or_404
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from apps.common.pagination import StandardResultsSetPagination
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from .serializers import AdminListSerializer, UserSerializer
from .permissions import IsAdmin
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from .filters import UserFilter
from .models import User
from .serializers import UserSerializer, RegisterSerializer

class TeacherOnlyUserListView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = StandardResultsSetPagination

    queryset = User.objects.exclude(phone_number='+998900748737').order_by('-created_at')

    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = UserFilter
    ordering_fields = ['first_name', 'last_name', 'email', 'created_at']
    ordering = ['-created_at']


class AdminUserListView(generics.ListAPIView):
    """
    GET /api/users/ -> List all users
    """
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = StandardResultsSetPagination
    queryset = User.objects.all().order_by('-created_at')
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = UserFilter

    def post(self, request):
        """
        POST /api/users/ -> Create a new user (simplified)
        """
        data = request.data
        # Only admins can create users
        if request.user.role != "admin":
            return Response({"detail": "Only admins can create users."}, status=status.HTTP_403_FORBIDDEN)

        telegram_username = data.get("telegram_username", "").strip().lstrip("@")
        existing = User.objects.filter(telegram_username=telegram_username).first()
        if existing:
            # Role duplication guard: if user already has a non-student role, reject
            if existing.role in ("teacher", "admin") and data.get("role") in ("teacher", "admin"):
                return Response(
                    {"detail": f"User @{telegram_username} already has the role '{existing.role}'. Cannot assign another privileged role."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        try:
            user = User.objects.create(
                telegram_username=telegram_username,
                username=telegram_username,
                full_name=data.get("full_name"),
                role=data.get("role", "student")
            )
            user.set_password(data.get("password"))
            user.save()
            return Response({"success": True}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class AdminCreateUserAPIView(generics.CreateAPIView):
    """
    Admin foydalanuvchi qo'shish endpointi
    POST /users/add/
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    queryset = User.objects.all()
    serializer_class = RegisterSerializer
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request, *args, **kwargs):
        # Form data'dagi barcha qiymatlarni to'g'ri qayta ishlash
        data = {}
        
        for key, value in request.data.items():
            if key == 'group':
                # group fieldini groups array'ga o'zgartirish
                if value:
                    if isinstance(value, list):
                        value = value[0] if value else None
                    if value:
                        data['groups'] = [int(value)]
            else:
                # Boshqa fieldlar uchun - agar list bo'lsa, birinchi elementni olish
                if isinstance(value, list):
                    data[key] = value[0] if value else None
                else:
                    data[key] = value
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()

        photo = request.FILES.get('photo')
        if photo:
            user.photo = photo
            user.save(update_fields=['photo'])

        return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)


class AdminUserUpdateView(generics.RetrieveUpdateDestroyAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    lookup_field = "pk"
    parser_classes = [MultiPartParser, FormParser, JSONParser]


class AdminChangeUserPasswordAPIView(APIView):
	"""
	Admin-only endpoint to change another user's password without old password.
	Student-only endpoint to change own password with old password.
	
	Admin case:
	POST /users/{user_id}/change-password/
	Body:
	{
	  "new_password": "newpass",
	  "new_confirm_password": "newpass"
	}
	
	Student case:
	POST /users/{user_id}/change-password/
	Body:
	{
	  "old_password": "oldpass",
	  "new_password": "newpass",
	  "new_confirm_password": "newpass"
	}
	"""
	permission_classes = [IsAuthenticated]

	def post(self, request, user_id, *args, **kwargs):
		current_user = request.user
		data = request.data or {}
		new_password = data.get("new_password")
		new_confirm_password = data.get("new_confirm_password")
		old_password = data.get("old_password")

		def _to_str(val):
			if isinstance(val, list):
				return val[0] if val else None
			if isinstance(val, bytes):
				try:
					return val.decode("utf-8")
				except Exception:
					return str(val)
			if val is None:
				return None
			return str(val)

		new_password = _to_str(new_password)
		new_confirm_password = _to_str(new_confirm_password)
		old_password = _to_str(old_password)

		target_user = get_object_or_404(User, id=user_id)
		errors = {}
		
		if current_user.role == 'student':
			if str(current_user.id) != str(user_id):
				return Response(
					{"detail": "Siz faqat o'zingizning parolingizni o'zgartira olasiz."},
					status=status.HTTP_403_FORBIDDEN
				)
			
			if not old_password:
				errors["old_password"] = ["Eski parolni jo'nating."]
			if not new_password:
				errors["new_password"] = ["Yangi parolni jo'nating."]
			if not new_confirm_password:
				errors["new_confirm_password"] = ["Parolni tasdiqlang."]
			
			if errors:
				return Response(errors, status=status.HTTP_400_BAD_REQUEST)
			
			if not current_user.check_password(old_password):
				return Response(
					{"old_password": ["Eski parol noto'g'ri."]},
					status=status.HTTP_400_BAD_REQUEST
				)
		
		elif current_user.role == 'admin':
			if not new_password:
				errors["new_password"] = ["buni jo'nating."]
			if not new_confirm_password:
				errors["new_confirm_password"] = ["buni jo'nating."]
			
			if errors:
				return Response(errors, status=status.HTTP_400_BAD_REQUEST)
		
		else:
			return Response(
				{"detail": "Sizda bu amalni bajarish huquqi yo'q."},
				status=status.HTTP_403_FORBIDDEN
			)

		if new_password != new_confirm_password:
			return Response(
				{"new_confirm_password": ["Parollar mos kelmadi."]},
				status=status.HTTP_400_BAD_REQUEST
			)

		target_user.set_password(new_password)
		target_user.save(update_fields=["password", "updated_at"])
		return Response({"detail": "Parol muvaffaqiyatli o'zgartirildi."}, status=status.HTTP_200_OK)


class AdminCheckUserAPIView(generics.ListAPIView):
    """
    Admin foydalanuvchi tekshirish endpointi
    GET /users/check/?uuid=...
    """
    permission_classes = [IsAuthenticated, IsAdmin]
    serializer_class = UserSerializer

    def get_queryset(self):
        uuid = self.request.query_params.get('uuid', None)
        if uuid is not None:
            return User.objects.filter(uuid=uuid)
        return User.objects.none()
    

class AdminsListView(generics.ListAPIView):
    """
    Admin foydalanuvchilar ro'yxati
    GET /users/admins/
    """
    permission_classes = [IsAuthenticated]
    serializer_class = AdminListSerializer

    def get_queryset(self):
        return User.objects.filter(role='admin').order_by('-created_at')


class UsersExportAPIView(APIView):
	"""Export all users as CSV. Admin only."""
	permission_classes = [IsAuthenticated]

	def get(self, request):
		if request.user.role != 'admin':
			return Response({"detail": "Sizda bu amalni bajarish huquqi yo'q."}, status=status.HTTP_403_FORBIDDEN)

		import csv
		from django.http import HttpResponse

		fieldnames = [
			'id', 'username', 'phone_number', 'first_name', 'last_name', 'email',
			'role', 'level', 'coins', 'created_at'
		]

		response = HttpResponse(content_type='text/csv')
		response['Content-Disposition'] = 'attachment; filename="users_export.csv"'

		writer = csv.DictWriter(response, fieldnames=fieldnames)
		writer.writeheader()
		for u in User.objects.all().order_by('-created_at'):
			writer.writerow({
				'id': str(u.id),
				'username': u.username,
				'phone_number': u.phone_number,
				'first_name': u.first_name or '',
				'last_name': u.last_name or '',
				'email': u.email or '',
				'role': u.role,
				'level': u.level or '',
				'coins': u.coins or 0,
				'created_at': u.created_at.isoformat() if u.created_at else '',
			})

		return response


class UsersTemplateAPIView(APIView):
	"""Return XLSX template (header only) for imports. Admin only."""
	permission_classes = [IsAuthenticated]

	def get(self, request):
		if request.user.role != 'admin':
			return Response({"detail": "Sizda bu amalni bajarish huquqi yo'q."}, status=status.HTTP_403_FORBIDDEN)

		from openpyxl import Workbook
		from django.http import HttpResponse
		import io

		# Create workbook
		wb = Workbook()
		ws = wb.active
		ws.title = "Talabalar"

		# Add headers
		fieldnames = ['Ism', 'Familya', 'Telefon_raqam', 'Ota', 'Ona']
		ws.append(fieldnames)

		# Style header row
		from openpyxl.styles import Font, PatternFill, Alignment
		fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
		font = Font(bold=True, color="FFFFFF")
		
		for cell in ws[1]:
			cell.fill = fill
			cell.font = font
			cell.alignment = Alignment(horizontal="center", vertical="center")

		# Set column widths
		ws.column_dimensions['A'].width = 15  # Ism
		ws.column_dimensions['B'].width = 15  # Familya
		ws.column_dimensions['C'].width = 18  # Telefon_raqam
		ws.column_dimensions['D'].width = 18  # Ota
		ws.column_dimensions['E'].width = 18  # Ona

		# Save to BytesIO
		output = io.BytesIO()
		wb.save(output)
		output.seek(0)

		response = HttpResponse(
			output.getvalue(),
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		response['Content-Disposition'] = 'attachment; filename="talabalar_shabloni.xlsx"'
		return response


class UsersImportAPIView(APIView):
	"""Import users from uploaded XLSX file. Admin only."""
	permission_classes = [IsAuthenticated]
	parser_classes = [MultiPartParser, FormParser]

	def post(self, request):
		if request.user.role != 'admin':
			return Response({"detail": "Sizda bu amalni bajarish huquqi yo'q."}, status=status.HTTP_403_FORBIDDEN)

		xlsx_file = request.FILES.get('file')
		if not xlsx_file:
			return Response({"detail": "XLSX fayl 'file' nomi bilan yuborilishi kerak."}, status=status.HTTP_400_BAD_REQUEST)

		from openpyxl import load_workbook
		import io

		created = []
		updated = []
		errors = []

		try:
			# Load workbook
			wb = load_workbook(io.BytesIO(xlsx_file.read()))
			ws = wb.active

			# Get headers
			headers = []
			for cell in ws[1]:
				headers.append(cell.value)

			# Process data rows (skip header)
			for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
				row_data = {}
				for col_num, cell in enumerate(row):
					if col_num < len(headers):
						# Convert to string va strip qil
						value = cell.value
						if value is None:
							row_data[headers[col_num]] = ''
						else:
							row_data[headers[col_num]] = str(value).strip()

				# Majburiy fieldlar
				first_name = row_data.get('Ism', '').strip()
				last_name = row_data.get('Familya', '').strip()
				phone_number = row_data.get('Telefon_raqam', '').strip()
				father_phone = row_data.get('Ota', '').strip()
				mother_phone = row_data.get('Ona', '').strip()

				# Validation
				if not phone_number or not first_name or not last_name:
					errors.append({
						"row": row_num,
						"detail": "Ism, Familya va Telefon_raqam majburiy."
					})
					continue

				if not (father_phone or mother_phone):
					errors.append({
						"row": row_num,
						"detail": "Ota yoki Ona dan biri majburiy."
					})
					continue

				try:
					username = phone_number
					
					# Parent phones JSONField uchun
					parent_phones = {}
					if father_phone:
						parent_phones['father'] = father_phone
					if mother_phone:
						parent_phones['mother'] = mother_phone
					
					# Check if user exists
					user = User.objects.filter(phone_number=phone_number).first()
					
					if user:
						# Update existing user
						user.first_name = first_name
						user.last_name = last_name
						user.username = username
						user.parent_phone_number = parent_phones if parent_phones else None
						user.save()
						updated.append(str(user.id))
					else:
						# Create new user
						user = User.objects.create(
							phone_number=phone_number,
							username=username,
							first_name=first_name,
							last_name=last_name,
							role='student',
							level='beginner',
							parent_phone_number=parent_phones if parent_phones else None,
							is_active=True
						)
						user.set_password(phone_number)
						user.save()
						created.append(str(user.id))
				except Exception as e:
					errors.append({"row": row_num, "detail": str(e)})

		except Exception as e:
			return Response(
				{"detail": f"XLSX faylni o'qishda xato: {str(e)}"}, 
				status=status.HTTP_400_BAD_REQUEST
			)

		return Response({"created": created, "updated": updated, "errors": errors}, status=status.HTTP_200_OK)
